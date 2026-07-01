import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


INPUTS = [
    Path("analysis_input/PURULIA TRUCK LOAD DETAILS (2026-27).xlsx"),
    Path("analysis_input/SHREE PURULIA PAYMENT (2026-27).xlsx"),
]


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def row_values(sheet, row_number, max_col=None):
    max_col = max_col or sheet.max_column
    return [clean(sheet.cell(row=row_number, column=column).value) for column in range(1, max_col + 1)]


def non_empty_rows(sheet):
    rows = []
    for row_number in range(1, sheet.max_row + 1):
        values = row_values(sheet, row_number)
        if any(values):
            rows.append((row_number, values))
    return rows


def style_id_row(sheet, row_number):
    return [sheet.cell(row=row_number, column=column).style_id for column in range(1, sheet.max_column + 1)]


def formula_cells(sheet):
    formulas = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append(
                    {
                        "cell": cell.coordinate,
                        "formula": cell.value,
                        "number_format": cell.number_format,
                    }
                )
    return formulas


def hidden_columns(sheet):
    return [
        column
        for column, dimension in sheet.column_dimensions.items()
        if dimension.hidden
    ]


def hidden_rows(sheet):
    return [
        row
        for row, dimension in sheet.row_dimensions.items()
        if dimension.hidden
    ]


def merged_ranges(sheet):
    return [str(item) for item in sheet.merged_cells.ranges]


def repeated_rows(rows):
    counter = Counter(tuple(values) for _, values in rows)
    return [
        {"count": count, "values": list(values)}
        for values, count in counter.most_common()
        if count > 1 and any(values)
    ][:20]


def likely_header_rows(rows):
    keywords = re.compile(
        r"(truck|vehicle|owner|pan|gst|date|challan|load|short|advance|balance|rate|amount|tds|freight|quantity|net)",
        re.I,
    )
    candidates = []
    for row_number, values in rows:
        score = sum(1 for value in values if keywords.search(value))
        filled = sum(1 for value in values if value)
        if score >= 2 or (score >= 1 and filled >= 4):
            candidates.append(
                {
                    "row": row_number,
                    "score": score,
                    "filledCells": filled,
                    "values": values,
                }
            )
    return candidates[:50]


def value_frequencies(sheet):
    frequencies = defaultdict(Counter)
    for row_number in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            value = clean(sheet.cell(row=row_number, column=column).value)
            if value:
                frequencies[get_column_letter(column)][value] += 1
    return {
        column: counter.most_common(10)
        for column, counter in frequencies.items()
        if counter
    }


def inspect_sheet(sheet):
    rows = non_empty_rows(sheet)
    formulas = formula_cells(sheet)
    first_rows = [{"row": row_number, "values": values} for row_number, values in rows[:30]]

    return {
        "title": sheet.title,
        "state": sheet.sheet_state,
        "dimensions": sheet.calculate_dimension(),
        "maxRow": sheet.max_row,
        "maxColumn": sheet.max_column,
        "hiddenRows": hidden_rows(sheet),
        "hiddenColumns": hidden_columns(sheet),
        "mergedRanges": merged_ranges(sheet),
        "formulaCount": len(formulas),
        "formulas": formulas[:100],
        "likelyHeaderRows": likely_header_rows(rows),
        "repeatedRows": repeated_rows(rows),
        "firstNonEmptyRows": first_rows,
        "columnWidths": {
            key: dimension.width
            for key, dimension in sheet.column_dimensions.items()
            if dimension.width
        },
        "rowHeights": {
            str(key): dimension.height
            for key, dimension in sheet.row_dimensions.items()
            if dimension.height
        },
        "valueFrequencies": value_frequencies(sheet),
    }


def inspect_workbook(path):
    workbook = load_workbook(path, data_only=False)
    return {
        "fileName": path.name,
        "sheetNames": workbook.sheetnames,
        "sheets": [inspect_sheet(workbook[sheet_name]) for sheet_name in workbook.sheetnames],
    }


def main():
    output_dir = Path("analysis_output")
    output_dir.mkdir(exist_ok=True)
    result = [inspect_workbook(path) for path in INPUTS]
    output_path = output_dir / "workbook_analysis.json"
    output_path.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
