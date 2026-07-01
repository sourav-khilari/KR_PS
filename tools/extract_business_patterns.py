import json
import re
from pathlib import Path

from openpyxl import load_workbook


FILES = [
    Path("analysis_input/PURULIA TRUCK LOAD DETAILS (2026-27).xlsx"),
    Path("analysis_input/SHREE PURULIA PAYMENT (2026-27).xlsx"),
]


LABEL_RE = re.compile(
    r"Truck Owner Name|PAN NO|PAYMENT SHEET|Total:|TAXABLE VALUE|ADD: CGST|ADD: SGST|NET BILL AMOUNT|LESS: DIESEL|LESS: CASH|LESS: TDS|NET PAYABLE|GST PAYABLE|Sl",
    re.I,
)


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def full_row(sheet, row_number):
    return [clean(sheet.cell(row=row_number, column=column).value) for column in range(1, sheet.max_column + 1)]


def extract_sheet(sheet):
    labels = []
    for row_number in range(1, sheet.max_row + 1):
        values = full_row(sheet, row_number)
        text = " ".join(values)
        if LABEL_RE.search(text):
            labels.append(
                {
                    "row": row_number,
                    "values": values,
                    "formulas": [
                        {
                            "cell": sheet.cell(row=row_number, column=column).coordinate,
                            "formula": sheet.cell(row=row_number, column=column).value,
                        }
                        for column in range(1, sheet.max_column + 1)
                        if isinstance(sheet.cell(row=row_number, column=column).value, str)
                        and sheet.cell(row=row_number, column=column).value.startswith("=")
                    ],
                }
            )
    return labels


def main():
    result = {}
    for path in FILES:
        workbook = load_workbook(path, data_only=False)
        result[path.name] = {sheet.title: extract_sheet(sheet) for sheet in workbook.worksheets}

    output = Path("analysis_output/business_patterns.json")
    output.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(output)


if __name__ == "__main__":
    main()
