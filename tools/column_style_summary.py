import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


FILES = [
    Path("analysis_input/PURULIA TRUCK LOAD DETAILS (2026-27).xlsx"),
    Path("analysis_input/SHREE PURULIA PAYMENT (2026-27).xlsx"),
]


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def cell_style(cell):
    fill = cell.fill.fgColor.rgb or cell.fill.fgColor.indexed or cell.fill.fgColor.theme
    return {
        "fontBold": cell.font.bold,
        "fontName": cell.font.name,
        "fontSize": cell.font.sz,
        "fill": str(fill),
        "borderStyle": {
            "left": cell.border.left.style,
            "right": cell.border.right.style,
            "top": cell.border.top.style,
            "bottom": cell.border.bottom.style,
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrapText": cell.alignment.wrap_text,
        },
        "numberFormat": cell.number_format,
    }


def header_row(sheet, row_number):
    return [
        {
            "column": get_column_letter(column),
            "value": clean(sheet.cell(row=row_number, column=column).value),
            "style": cell_style(sheet.cell(row=row_number, column=column)),
            "width": sheet.column_dimensions[get_column_letter(column)].width,
        }
        for column in range(1, sheet.max_column + 1)
    ]


def main():
    result = {}
    for path in FILES:
        workbook = load_workbook(path, data_only=False)
        result[path.name] = {}
        for sheet in workbook.worksheets:
            rows = []
            for row_number in range(1, min(sheet.max_row, 200) + 1):
                values = [clean(sheet.cell(row=row_number, column=column).value) for column in range(1, sheet.max_column + 1)]
                if "Truck No" in values or "TRUCK NO." in values:
                    rows.append({"row": row_number, "columns": header_row(sheet, row_number)})
            result[path.name][sheet.title] = rows

    output = Path("analysis_output/column_style_summary.json")
    output.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(output)


if __name__ == "__main__":
    main()
